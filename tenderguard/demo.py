from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


def demo_boq_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOQ"
    headers = ["Item No.", "Section", "Description", "Unit", "Quantity", "Rate", "Amount"]
    rows = [
        ["D001", "Waterproofing", "Waterproof membrane to toilet floor", "m2", 50, 180, 9000],
        ["D002", "Waterproofing", "Waterproof membrane to toilet walls", "m2", 85, 195, 16000],
        ["D003", "Sealants", "Sealant to movement joints", "m", 200, 0, 0],
        ["D004", "Waterproofing", "Waterproof membrane to toilet floor", "m2", 50, 180, 9000],
        ["D005", "Doors", "Timber door set", "m2", 12, 1200, 14400],
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="16324F")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    widths = [13, 20, 46, 10, 12, 14, 15]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=2, min_col=5, max_col=7):
        for cell in row:
            cell.number_format = "#,##0.00"
    sheet.freeze_panes = "A2"
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def demo_specification_bytes() -> bytes:
    stream = io.BytesIO()
    pdf = canvas.Canvas(stream, pagesize=A4)
    width, height = A4

    def line(text: str, y: float, size: int = 11, bold: bool = False) -> float:
        pdf.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        pdf.drawString(62, y, text)
        return y - (size + 10)

    y = height - 70
    y = line("Sample Tender Specification", y, 18, True)
    y -= 12
    y = line("Section 12 - Waterproofing", y, 14, True)
    y -= 8
    y = line("Clause 12.1", y, 11, True)
    y = line("Provide waterproof membrane to toilet floors, including preparation and testing.", y)
    y -= 8
    y = line("Clause 12.2", y, 11, True)
    y = line("Provide waterproof membrane to toilet walls.", y)
    y -= 8
    y = line("Clause 12.3", y, 11, True)
    y = line("Provide waterproof membrane to all balcony floors, including upstands and", y)
    y = line("sealing around penetrations.", y)
    pdf.setFont("Helvetica", 9)
    pdf.drawRightString(width - 50, 35, "Page 1")
    pdf.showPage()

    y = height - 70
    y = line("Section 13 - Sealants", y, 14, True)
    y -= 8
    y = line("Clause 13.1", y, 11, True)
    y = line("Provide sealant to movement joints where indicated.", y)
    y -= 24
    y = line("Section 14 - Doors", y, 14, True)
    y -= 8
    y = line("Clause 14.1", y, 11, True)
    y = line("Provide complete timber door sets, including frames and ironmongery.", y)
    pdf.setFont("Helvetica", 9)
    pdf.drawRightString(width - 50, 35, "Page 2")
    pdf.save()
    return stream.getvalue()

